import os
import openpyxl
from time import sleep
from mast.cli import Cli
from mast.logging import make_logger
from mast.datapower import datapower
from mast.timestamp import Timestamp
from mast.datapower.datapower import Environment

t = Timestamp()

logger = make_logger("mast.getstatus", propagate=False)


def add_to_header(column_name, header_row):
    if column_name in header_row:
        pass
    else:
        header_row.append(column_name)


def _recurse_config(prefix, node, row, header_row, delim, obfuscate_password):
    if obfuscate_password and node.tag.lower() == "password":
        # Obfuscate any passwords
        node.text = "XXXXXXXXXX"
    if list(node):
        for child in list(node):
            new_prefix = "{}/{}".format(prefix, child.tag)
            logger.info("Found child node {}".format(new_prefix))
            _recurse_config(
                new_prefix,
                child,
                row,
                header_row,
                delim,
                obfuscate_password)
        for attr in node.keys():
            value = node.get(attr).strip() or "N/A"
            field_name = "{}[@{}]".format(prefix, attr)
            logger.info("Found attribute {} - {}".format(field_name, value))
            add_to_header(field_name, header_row)
            row[header_row.index(field_name)] = value
    else:
        logger.info("No child nodes found")
        if not node.text:
            node.text = "N/A"
        add_to_header(prefix, header_row)
        index = header_row.index(prefix)
        if not row[index]:
            logger.info(
                "Found value for {} -> {}".format(
                    node.tag.encode("utf-8"),
                    node.text.encode("utf-8")
                )
            )
            row[index] = node.text
        else:
            logger.info(
                "Found additional value for {} -> {}".format(
                    node.tag.encode("utf-8"),
                    node.text.encode("utf-8")
                )
            )
            row[index] = "{}{}{}".format(row[index], delim, node.text)
        for attr in node.keys():
            value = node.get(attr).strip() or "N/A"
            field_name = "{}[@{}]".format(prefix, attr)
            logger.info("Found attribute {} -> {}".format(field_name, value))
            add_to_header(field_name, header_row)
            row[header_row.index(field_name)] = value


def recurse_status(prefix, elem, row, header_row):
    for child in elem.findall(".//*"):
        logger.info("Found child {}".format(child.tag))
        if child.tag not in header_row:
            logger.debug("Appending {} to header row".format(child.tag))
            header_row.append(child.tag)
        if len(child):
            _prefix = "{}/{}".format(prefix, child.tag)
            logger.info("child {} has children, recursing...".format(_prefix))
            recurse_status(_prefix, child, row, header_row)
        else:
            logger.info("Adding {} -> {}".format(child.tag, child.text))
            row.insert(header_row.index(child.tag), child.text)


def create_workbook(
        env,       domains,
        providers, object_classes, delay,
        delim, out_file,  timestamp,
        prepend_timestamp, obfuscate_password):
    if not providers:
        providers = []
    if not object_classes:
        object_classes = []

    xpath = datapower.STATUS_XPATH
    wb = openpyxl.Workbook()
    base_header_row = ["timestamp", "hostname", "domain", "provider"]

    if prepend_timestamp:
        filename = os.path.split(out_file)[-1]
        filename = "{}-{}".format(t.timestamp, filename)
        path = list(os.path.split(out_file)[:-1])
        path.append(filename)
        out_file = os.path.join(*path)
    path = os.path.join(*list(os.path.split(out_file)[:-1]))
    if not os.path.exists(path):
        os.makedirs(path)

    logger.info("Generating Workbook - will be stored in {}".format(out_file))
    skip = []
    for provider in providers:
        sheet = wb.create_sheet(title=provider[:31])
        header_row = list(base_header_row)
        rows = []
        for appliance in env.appliances:
            if appliance in skip:
                continue
            print "checking {}".format(provider)
            logger.info("Checking Status Provider {}".format(provider))
            print "\t{}".format(appliance.hostname)
            try:
                if "all-domains" in domains:
                    _domains = appliance.domains
                else:
                    _domains = domains
            except:
                print " ".join((
                    "ERROR: See log for details,",
                    "skipping appliance {}".format(appliance.hostname)
                ))
                logger.exception(
                    "An unhandled exception was raised while "
                    "retrieving list of domains."
                    "Skipping appliance {}.".format(appliance.hostname)
                )
                skip.append(appliance)
                continue
            for domain in _domains:
                print "\t\t{}".format(domain)
                sleep(delay)
                try:
                    logger.info(
                        "Querying {} for {} in domain {}".format(
                            appliance.hostname,
                            provider,
                            domain
                        )
                    )
                    response = appliance.get_status(provider, domain=domain)
                except datapower.AuthenticationFailure:
                    logger.warn(
                        "Recieved AuthenticationFailure. "
                        "Retrying in 5 seconds...")
                    print " ".join((
                        "Recieved AuthenticationFailure.",
                        "Retrying in 5 seconds..."
                    ))
                    sleep(5)
                    try:
                        response = appliance.get_status(
                            provider,
                            domain=domain
                        )
                    except datapower.AuthenticationFailure:
                        print "Received AuthenticationFailure again. Skipping."
                        logger.error(
                            "Received AuthenticationFailure again. Skipping."
                        )
                        skip.append(appliance)
                        continue
                except:
                    print " ".join((
                        "ERROR: See log for details,",
                        "skipping appliance {}".format(appliance.hostname)
                    ))
                    logger.exception(
                        "An unhandled exception was raised. "
                        "Skipping appliance {}.".format(appliance.hostname)
                    )
                    skip.append(appliance)
                    break
                timestamp = response.xml.find(datapower.BASE_XPATH).text
                for node in response.xml.findall(xpath):
                    row = [timestamp, appliance.hostname, domain, provider]
                    recurse_status("", node, row, header_row)
                    rows.append(row)
        sheet.append(header_row)
        for row in rows:
            sheet.append(row)

    for object_class in object_classes:
        header_row = ["appliance", "domain", "Object Class", "Object Name"]
        rows = []
        ws = wb.create_sheet(title=object_class)
        for dp in env.appliances:
            if dp in skip:
                continue
            print "Retrieving {}".format(object_class)
            logger.info("Retrieving {} configuration".format(object_class))
            print "\t{}".format(dp.hostname)
            logger.info("Querying {}".format(dp.hostname))
            _domains = domains
            if "all-domains" in domains:
                try:
                    _domains = dp.domains
                except:
                    print " ".join((
                        "ERROR: See log for details,",
                        "skipping appliance {}".format(dp.hostname)
                    ))
                    logger.exception(
                        " ".join((
                            "An unhandled exception was raised",
                            "while retrieving list of domains.",
                            "Skipping appliance {}.".format(dp.hostname)
                        ))
                    )
                    skip.append(dp)
                    continue
            for domain in _domains:
                print "\t\t{}".format(domain)
                logger.info("Looking in domain {}".format(domain))
                xpath = datapower.CONFIG_XPATH + object_class
                try:
                    logger.info(
                        "Querying {} for {} in domain {}".format(
                            dp.hostname,
                            object_class,
                            domain
                        )
                    )
                    config = dp.get_config(_class=object_class, domain=domain)
                except datapower.AuthenticationFailure:
                    logger.warn(
                        "Recieved AuthenticationFailure."
                        "Retrying in 5 seconds...")
                    print " ".join((
                        "Recieved AuthenticationFailure.",
                        "Retrying in 5 seconds..."
                    ))
                    sleep(5)
                    try:
                        config = dp.get_config(
                            _class=object_class,
                            domain=domain
                        )
                    except datapower.AuthenticationFailure:
                        print "Received AuthenticationFailure again. Skipping."
                        logger.error(
                            "Received AuthenticationFailure again. Skipping.")
                        skip.append(dp)
                        continue
                except:
                    print " ".join((
                        "ERROR: See log for details,",
                        "skipping appliance {}".format(dp.hostname)
                    ))
                    logger.exception(
                        " ".join((
                            "An unhandled exception was raised.",
                            "Skipping appliance {}.".format(dp.hostname)
                        ))
                    )
                    skip.append(dp)
                    break
                nodes = config.xml.findall(xpath)
                for index, node in enumerate(nodes):
                    name = node.get("name")
                    logger.info("Found node {} - {}".format(node.tag, name))
                    row = [dp.hostname, domain, object_class, name]
                    row.extend([None] * 1500)
                    for child in list(node):
                        logger.info(
                            "Found child node {}. recursing...".format(child)
                        )
                        _recurse_config(
                            child.tag,
                            child,
                            row,
                            header_row,
                            delim,
                            obfuscate_password
                        )
                    rows.append(row)
        rows.insert(0, header_row)
        for row in rows:
            ws.append(row)


    Sheet = wb.get_sheet_by_name("Sheet")
    wb.remove_sheet(Sheet)
    logger.info("Writing workbook {}".format(out_file))
    wb.save(out_file)


def main(
        appliances=[], credentials=[],
        domains=["default"],      providers=[],
        object_classes=[],        delim=",",
        timeout=120,              delay=0.5,
        out_file="./status.xlsx", no_check_hostname=False,
        by_appliance=False,       no_prepend_timestamp=False,
        obfuscate_password=False):

    prepend_timestamp = not no_prepend_timestamp
    t = Timestamp()

    check_hostname = not no_check_hostname
    if by_appliance:
        logger.info("Generating workbooks by appliance")
        # we make the initial Environment to correctly handle credentials
        env = Environment(
            appliances,
            credentials,
            timeout,
            check_hostname=check_hostname
        )
        for appliance in env.appliances:
            logger.info(
                "generating workbook for {}".format(appliance.hostname)
            )
            filename = os.path.split(out_file)[-1]
            filename = "{}-{}".format(appliance.hostname, filename)
            path = list(os.path.split(out_file)[:-1])
            path.append(filename)
            _out_file = os.path.join(*path)
            logger.info(
                "Workbook for {} will be stored in {}".format(
                    appliance.hostname, _out_file
                )
            )
            _env = Environment(
                [appliance.hostname],
                [appliance.credentials],
                timeout,
                check_hostname)

            create_workbook(
                _env,
                domains,
                providers,
                object_classes,
                delay,
                delim,
                _out_file,
                t,
                prepend_timestamp,
                obfuscate_password
            )
    else:
        logger.info("generating workbook")
        env = Environment(
            appliances,
            credentials,
            timeout,
            check_hostname=False
        )
        create_workbook(
            env,
            domains,
            providers,
            object_classes,
            delay,
            delim,
            out_file,
            t,
            prepend_timestamp,
            obfuscate_password
        )


if __name__ == "__main__":
    try:
        cli = Cli(main=main)
        cli.run()
    except:
        logger.exception("An unhandled exception occured during execution.")
        raise
